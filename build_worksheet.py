# -*- coding: utf-8 -*-
"""NLPコーチング ワークシート（代表システム優先度テスト＋望むこと100個）を1つのxlsxに統合"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import RadarChart, Reference

# ===== アプリと統一したカラーパレット =====
BRAND   = "2f8f7f"  # メインのティールグリーン
BRAND_D = "226b5f"
BRAND_L = "e7f4f1"
ACCENT  = "6b6ad6"
ACCENT_L= "ececfb"
WARN    = "c98a2b"
WARN_L  = "fbf2e3"
INK     = "23323a"
MUTED   = "6b7c85"
LINE    = "dbe6e3"
PALE    = "f4f7f6"
WHITE   = "ffffff"

# ===== 共通スタイル =====
def font(size=11, bold=False, color=INK, name="游ゴシック"):
    return Font(name=name, size=size, bold=bold, color=color)

def fill(c): return PatternFill("solid", fgColor=c)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

thin_border = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)
medium_brand = Border(
    left=Side(style="medium", color=BRAND),
    right=Side(style="medium", color=BRAND),
    top=Side(style="medium", color=BRAND),
    bottom=Side(style="medium", color=BRAND),
)

# ===========================================================
def build():
    wb = Workbook()
    # デフォルトシート削除
    wb.remove(wb.active)

    cover_sheet(wb)
    rep_test_sheet(wb)
    rep_type_sheet(wb)
    balance_wheel_sheet(wb)
    themes_sheet(wb)
    list100_sheet(wb)

    out = "nlp_worksheet.xlsx"
    wb.save(out)
    print(f"{out} を生成しました")


# ===========================================================
# 1) 表紙
# ===========================================================
def cover_sheet(wb):
    ws = wb.create_sheet("はじめに")
    ws.sheet_view.showGridLines = False

    # 全体の背景白
    for r in range(1, 50):
        for c in range(1, 12):
            ws.cell(row=r, column=c).fill = fill(WHITE)

    # ヘッダー帯
    ws.row_dimensions[1].height = 4
    for c in range(1, 12):
        ws.cell(row=1, column=c).fill = fill(BRAND)
    # タイトル帯
    ws.row_dimensions[3].height = 60
    ws.merge_cells("B3:K3")
    cell = ws["B3"]
    cell.value = "NLPコーチング ワークシート"
    cell.font = font(28, True, BRAND_D)
    cell.alignment = align("center", "center", False)

    ws.row_dimensions[4].height = 24
    ws.merge_cells("B4:K4")
    sub = ws["B4"]
    sub.value = "代表システム優先度テスト ＋ 望むこと100個 リスト"
    sub.font = font(13, False, MUTED)
    sub.alignment = align("center", "center", False)

    # 線
    ws.row_dimensions[5].height = 4
    for c in range(2, 12):
        ws.cell(row=5, column=c).fill = fill(BRAND_L)

    # ── 使い方 ──
    ws.row_dimensions[7].height = 26
    ws.merge_cells("B7:K7")
    h = ws["B7"]
    h.value = "📘 このワークシートの使い方"
    h.font = font(14, True, BRAND_D)
    h.alignment = align("left", "center", False)

    info = [
        ("① 代表システム優先度テスト",
         "あなたが世界を捉えるとき、視覚(V) / 聴覚(A) / 身体感覚(K) / 内部対話(Ad) のどの感覚を優位に使っているかを診断します。各設問の4つの選択肢に 4・3・2・1 を1回ずつ入力してください（同点不可）。"),
        ("② バランス車輪",
         "人生の8つの領域について、現在の満足度を 0〜10 で自己評価します。人生全体のバランスを可視化するステップです。"),
        ("③ テーマ別 望むこと",
         "各テーマごとに「将来どうなりたいか」「望むこと」を書き出します。記入した内容は自動的に「望むこと100個」シートに集約されます。"),
        ("④ 望むこと100個",
         "テーマ別シートに入力した内容が、100個のリストとして自動で集約・表示されます。人生全体の理想像の一覧です。"),
    ]
    r = 9
    for label, body in info:
        ws.row_dimensions[r].height = 22
        ws.merge_cells(f"B{r}:K{r}")
        cell = ws.cell(row=r, column=2)
        cell.value = label
        cell.font = font(12, True, BRAND_D)
        cell.alignment = align("left", "center", False)
        cell.fill = fill(BRAND_L)
        cell.border = thin_border

        ws.row_dimensions[r+1].height = 46
        ws.merge_cells(f"B{r+1}:K{r+1}")
        b = ws.cell(row=r+1, column=2)
        b.value = body
        b.font = font(11, False, INK)
        b.alignment = align("left", "top", True)
        b.fill = fill(PALE)
        b.border = thin_border
        r += 3

    # 注意書き
    r += 1
    ws.row_dimensions[r].height = 60
    ws.merge_cells(f"B{r}:K{r}")
    note = ws.cell(row=r, column=2)
    note.value = ("📌 進め方のヒント\n"
                  "・順番にこだわらず、書きやすいテーマから始めて構いません。\n"
                  "・「すでにそうなった」気持ちで、過去形・断定形で書くと潜在意識に届きやすくなります。\n"
                  "・1テーマあたり平均6〜8個を目安に。100個に届かなくても大丈夫です、思いついた数で進めましょう。")
    note.font = font(11, False, INK)
    note.alignment = align("left", "top", True)
    note.fill = fill(WARN_L)
    note.border = thin_border

    # 列幅
    ws.column_dimensions["A"].width = 2
    for c in "BCDEFGHIJK":
        ws.column_dimensions[c].width = 11
    ws.column_dimensions["L"].width = 2


# ===========================================================
# 2) 代表システム優先度テスト
# ===========================================================
QUESTIONS = [
    ("私は、大切な決定を＿＿＿を基に決めています。", [
        ("そのとき受けている感じの具合", "K"),
        ("どれが一番よい響きに聞こえるか", "A"),
        ("どれが最もよいビジョンが描けるか", "V"),
        ("正確に調べテーマを吟味すること", "Ad"),
    ]),
    ("議論の間、私が最も影響を受けやすいのは", [
        ("他の人が話す時の声の大きさや調子", "A"),
        ("他の人の話の概要がハッキリと見えるかどうか", "V"),
        ("他の人の話の筋道や論理", "Ad"),
        ("他の人の気持ちや内容に共感できるかどうか", "K"),
    ]),
    ("私が自分の状態を一番気づきやすいのは", [
        ("自分がどんな服を着て、見栄えはどうかによって", "V"),
        ("自分がどんな雰囲気になっているかによって", "K"),
        ("自分が話の中で意識する言葉づかいや内容によって", "Ad"),
        ("自分が話す時の声の調子によって", "A"),
    ]),
    ("私にとって、最も簡単にできそうなのは", [
        ("ステレオなどで好みの音質を見つけたりテレビの音量を気にすること", "A"),
        ("興味ある事柄に関連した知識や情報を集めること", "Ad"),
        ("一番快適で座り心地の良いソファーを探し出すこと", "K"),
        ("美しく、魅力的な色の配色を選び出すこと", "V"),
    ]),
    ("私は・・・", [
        ("私は自分のいる場所の音やざわめきが気になります", "A"),
        ("私は新しい情報やデータをよく知っていることに関心があります", "Ad"),
        ("私は自分に触れる椅子や布団の肌触りが気になります", "K"),
        ("私は自分の部屋の片づけ具合や色の配色に関心があります", "V"),
    ]),
]
TYPE_LABEL = {"V":"V（Visual）視覚型", "A":"A（Auditory）聴覚型",
              "K":"K（Kinesthetic）身体感覚型", "Ad":"Ad（Auditory Digital）内部対話型"}

def rep_test_sheet(wb):
    ws = wb.create_sheet("①代表システム優先度テスト")
    ws.sheet_view.showGridLines = False

    # タイトル帯
    ws.row_dimensions[1].height = 6
    for c in range(1, 8):
        ws.cell(row=1, column=c).fill = fill(BRAND)
    ws.row_dimensions[2].height = 38
    ws.merge_cells("B2:F2")
    t = ws["B2"]
    t.value = "代表システム 優先度テスト"
    t.font = font(20, True, BRAND_D)
    t.alignment = align("left", "center", False)

    # 説明
    ws.row_dimensions[3].height = 36
    ws.merge_cells("B3:F3")
    d = ws["B3"]
    d.value = ("各設問の4つの選択肢に対し、あなたに当てはまる度合いで「4 / 3 / 2 / 1」をそれぞれ1回ずつ入力してください（同点不可）。\n"
               "4=もっともあてはまる ／ 3=あてはまる ／ 2=ややあてはまる ／ 1=すこしあてはまる")
    d.font = font(10.5, False, MUTED)
    d.alignment = align("left", "top", True)

    # ヘッダー行
    headers = ["", "設問", "選択肢", "得点（1〜4）", "タイプ"]
    r = 5
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=i)
        cell.value = h
        cell.font = font(11, True, WHITE)
        cell.fill = fill(BRAND)
        cell.alignment = align("center", "center", False)
        cell.border = thin_border
    ws.row_dimensions[r].height = 26

    # データ検証（1〜4の整数のみ）
    dv = DataValidation(type="whole", operator="between", formula1=1, formula2=4,
                       errorTitle="入力エラー", error="1〜4の整数を入力してください",
                       allow_blank=True)
    dv.add(f"D{r+1}:D{r+1+25}")  # 5問×4選択肢=20行 + 余裕
    ws.add_data_validation(dv)

    # 設問行を生成（各問題の最初の行にBに問題、4行の選択肢）
    r += 1
    score_cells = {"V":[], "A":[], "K":[], "Ad":[]}
    for qi, (qtext, choices) in enumerate(QUESTIONS, 1):
        # 問題テキスト行（4行マージ）
        ws.merge_cells(start_row=r, start_column=2, end_row=r+3, end_column=2)
        qc = ws.cell(row=r, column=2)
        qc.value = f"問{qi}\n{qtext}"
        qc.font = font(11, True, BRAND_D)
        qc.alignment = align("left", "top", True)
        qc.fill = fill(BRAND_L)
        # 番号列
        ws.merge_cells(start_row=r, start_column=1, end_row=r+3, end_column=1)
        nc = ws.cell(row=r, column=1)
        nc.value = qi
        nc.font = font(18, True, BRAND)
        nc.alignment = align("center", "center", False)
        nc.fill = fill(BRAND_L)

        for ci, (ctxt, ctype) in enumerate(choices):
            row = r + ci
            # 選択肢
            cc = ws.cell(row=row, column=3)
            cc.value = ctxt
            cc.font = font(10.5, False, INK)
            cc.alignment = align("left", "center", True)
            cc.fill = fill(PALE if (qi % 2 == 1) else WHITE)
            cc.border = thin_border
            # 得点入力欄
            sc = ws.cell(row=row, column=4)
            sc.font = font(13, True, BRAND_D)
            sc.alignment = align("center", "center", False)
            sc.fill = fill(WHITE)
            sc.border = medium_brand
            # タイプ列
            tc = ws.cell(row=row, column=5)
            tc.value = ctype
            tc.font = font(11, True, ACCENT)
            tc.alignment = align("center", "center", False)
            tc.fill = fill(PALE if (qi % 2 == 1) else WHITE)
            tc.border = thin_border
            # cellの座標を保存
            score_cells[ctype].append(f"D{row}")
            # その他のセル装飾
            for col in (2,):
                pass
            ws.row_dimensions[row].height = 28
        r += 4
        # 区切り行
        ws.row_dimensions[r].height = 6
        r += 1

    # 自動集計テーブル
    r += 1
    ws.row_dimensions[r].height = 30
    ws.merge_cells(f"B{r}:E{r}")
    tt = ws.cell(row=r, column=2)
    tt.value = "📊 タイプ別 集計（自動計算）"
    tt.font = font(14, True, BRAND_D)
    tt.alignment = align("left", "center", False)
    r += 1
    headers2 = ["", "タイプ", "合計点", "比率"]
    for i, h in enumerate(headers2):
        c = ws.cell(row=r, column=i+2)
        c.value = h
        c.font = font(11, True, WHITE)
        c.fill = fill(BRAND_D)
        c.alignment = align("center", "center", False)
        c.border = thin_border
    ws.row_dimensions[r].height = 24
    r += 1
    start_total_row = r
    for tkey in ["V", "A", "K", "Ad"]:
        # ラベル列
        ws.cell(row=r, column=2).value = ""
        nc = ws.cell(row=r, column=3)
        nc.value = TYPE_LABEL[tkey]
        nc.font = font(11, True, INK)
        nc.alignment = align("left", "center", False)
        nc.fill = fill(BRAND_L)
        nc.border = thin_border
        # 合計
        sc = ws.cell(row=r, column=4)
        formula = "=SUM(" + ",".join(score_cells[tkey]) + ")"
        sc.value = formula
        sc.font = font(13, True, BRAND_D)
        sc.alignment = align("center", "center", False)
        sc.border = thin_border
        # 比率
        # 全体合計は4タイプの合計
        rc = ws.cell(row=r, column=5)
        rc.value = f"=IF(SUM($D${start_total_row}:$D${start_total_row+3})=0,0,D{r}/SUM($D${start_total_row}:$D${start_total_row+3}))"
        rc.number_format = "0%"
        rc.font = font(11, False, MUTED)
        rc.alignment = align("center", "center", False)
        rc.border = thin_border
        ws.row_dimensions[r].height = 26
        r += 1

    # 条件付き書式（合計点に応じてグラデーション）
    rule = ColorScaleRule(
        start_type='min', start_color=BRAND_L,
        end_type='max', end_color=BRAND
    )
    ws.conditional_formatting.add(f"D{start_total_row}:D{start_total_row+3}", rule)

    # 最優位タイプ
    r += 1
    ws.row_dimensions[r].height = 36
    ws.merge_cells(f"B{r}:C{r}")
    top_label = ws.cell(row=r, column=2)
    top_label.value = "🏆 最優位タイプ"
    top_label.font = font(13, True, ACCENT)
    top_label.alignment = align("left", "center", False)
    top_label.fill = fill(ACCENT_L)
    top_label.border = thin_border

    ws.merge_cells(f"D{r}:E{r}")
    top_val = ws.cell(row=r, column=4)
    # 最大値のタイプを返す数式
    top_val.value = (
        f'=IF(SUM(D{start_total_row}:D{start_total_row+3})=0,"未入力",'
        f'INDEX($C${start_total_row}:$C${start_total_row+3},'
        f'MATCH(MAX($D${start_total_row}:$D${start_total_row+3}),$D${start_total_row}:$D${start_total_row+3},0)))'
    )
    top_val.font = font(14, True, BRAND_D)
    top_val.alignment = align("center", "center", False)
    top_val.fill = fill(WHITE)
    top_val.border = medium_brand

    # 列幅
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 56
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 4


# ===========================================================
# 3) 優位感覚別 特徴
# ===========================================================
TYPE_FEATURES = [
    ("V（Visual）視覚型", BRAND, [
        "一般に背筋が伸びた良い姿勢",
        "目がよく上に動く",
        "話すテンポが速い",
        "声のトーンが高い",
        "話を視覚的な表現で好む",
        "外見を重視する",
        "呼吸は胸の上部で浅め",
    ]),
    ("A（Auditory）聴覚型", ACCENT, [
        "話すテンポはVとKの中間",
        "視線は左右によく動かす",
        "言葉の音を大切にする",
        "頭を傾ける",
        "声は澄んで歯切れよく響く",
        "音楽を聴いたり電話で話すのが好き",
        "騒音があると集中できない",
    ]),
    ("K（Kinesthetic）身体感覚型", WARN, [
        "お腹でゆっくり呼吸",
        "筋肉は緩んでいる",
        "頭は垂れている",
        "視線は下の方に動かす",
        "声は低い",
        "ゆっくりと間を置く話し方",
        "感触や触れ合いを大切にする",
        "人の近くに立つ傾向がある",
        "身体を動かしながらものを覚える",
    ]),
    ("Ad（Auditory Digital）内部対話型", BRAND_D, [
        "言葉の論理性を重視する",
        "落ち着いた話し方をする",
        "複雑な文を筋道立てて話すことができる",
        "独り言を言ったり自分の中で対話する",
    ]),
]

def rep_type_sheet(wb):
    ws = wb.create_sheet("②優位感覚別 特徴")
    ws.sheet_view.showGridLines = False

    # ヘッダー帯
    ws.row_dimensions[1].height = 6
    for c in range(1, 8):
        ws.cell(row=1, column=c).fill = fill(BRAND)
    ws.row_dimensions[2].height = 38
    ws.merge_cells("B2:F2")
    t = ws["B2"]
    t.value = "優位感覚別の特徴"
    t.font = font(20, True, BRAND_D)
    t.alignment = align("left", "center", False)

    ws.row_dimensions[3].height = 28
    ws.merge_cells("B3:F3")
    d = ws["B3"]
    d.value = "あなたの優位感覚を知ることで、自分の強みや、コーチング・対話で活かす切り口が見えてきます。"
    d.font = font(10.5, False, MUTED)
    d.alignment = align("left", "center", True)

    r = 5
    for tlabel, tcolor, feats in TYPE_FEATURES:
        # タイプ見出し
        ws.row_dimensions[r].height = 32
        ws.merge_cells(f"B{r}:F{r}")
        cell = ws.cell(row=r, column=2)
        cell.value = tlabel
        cell.font = font(14, True, WHITE)
        cell.fill = fill(tcolor)
        cell.alignment = align("left", "center", False)
        r += 1
        for f in feats:
            ws.row_dimensions[r].height = 22
            ws.merge_cells(f"B{r}:F{r}")
            cc = ws.cell(row=r, column=2)
            cc.value = "・ " + f
            cc.font = font(11, False, INK)
            cc.alignment = align("left", "center", True)
            cc.fill = fill(PALE)
            cc.border = thin_border
            r += 1
        # 区切り
        ws.row_dimensions[r].height = 8
        r += 1

    ws.column_dimensions["A"].width = 4
    for c in "BCDEFG":
        ws.column_dimensions[c].width = 18


# ===========================================================
# 4) バランス車輪
# ===========================================================
WHEEL = [
    ("仕事・キャリア", ["やりがい", "収入"]),
    ("お金・経済",    ["貯蓄", "投資"]),
    ("健康",          ["身体", "メンタル"]),
    ("家族・パートナー", ["家族", "パートナー"]),
    ("人間関係",      ["友人・知人", "仕事関係"]),
    ("精神性",        ["ミッション", "学び"]),
    ("遊び・余暇",    ["行きたい所", "欲しいもの"]),
    ("環境",          ["住居", "環境"]),
]

def balance_wheel_sheet(wb):
    ws = wb.create_sheet("③バランス車輪")
    ws.sheet_view.showGridLines = False

    # タイトル帯
    ws.row_dimensions[1].height = 6
    for c in range(1, 10):
        ws.cell(row=1, column=c).fill = fill(BRAND)
    ws.row_dimensions[2].height = 38
    ws.merge_cells("B2:G2")
    t = ws["B2"]
    t.value = "人生のバランス車輪"
    t.font = font(20, True, BRAND_D)
    t.alignment = align("left", "center", False)

    # 説明
    ws.row_dimensions[3].height = 56
    ws.merge_cells("B3:G3")
    d = ws["B3"]
    d.value = ("・人生の8つの領域のバランスを見ます。「0〜10」で満足度を自己評価してください。\n"
               "・現在の状態を記入してOKです。1ヶ月後・1年後・3年後など、将来の目標値としても活用できます。")
    d.font = font(10.5, False, MUTED)
    d.alignment = align("left", "top", True)

    # テーブル ヘッダー
    headers = ["", "大テーマ", "平均値", "小テーマ", "満足度（0〜10）"]
    r = 5
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=i+2)
        c.value = h
        c.font = font(11, True, WHITE)
        c.fill = fill(BRAND)
        c.alignment = align("center", "center", False)
        c.border = thin_border
    ws.row_dimensions[r].height = 28

    # データ検証（0〜10の整数）
    dv = DataValidation(type="whole", operator="between", formula1=0, formula2=10,
                       errorTitle="入力エラー", error="0〜10の整数を入力してください",
                       allow_blank=True)
    ws.add_data_validation(dv)

    r += 1
    start_data_row = r
    for i, (big, smalls) in enumerate(WHEEL):
        # 大テーマ（2行マージ）
        ws.merge_cells(start_row=r, start_column=3, end_row=r+1, end_column=3)
        bc = ws.cell(row=r, column=3)
        bc.value = big
        bc.font = font(12, True, BRAND_D)
        bc.alignment = align("left", "center", True)
        bc.fill = fill(BRAND_L if i % 2 == 0 else PALE)
        bc.border = thin_border
        # 平均値（2行マージ・自動）
        ws.merge_cells(start_row=r, start_column=4, end_row=r+1, end_column=4)
        avg = ws.cell(row=r, column=4)
        avg.value = f"=IFERROR(AVERAGE(F{r}:F{r+1}),\"\")"
        avg.font = font(13, True, ACCENT)
        avg.alignment = align("center", "center", False)
        avg.fill = fill(ACCENT_L)
        avg.border = thin_border
        avg.number_format = "0.0"
        # 番号
        ws.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=2)
        nc = ws.cell(row=r, column=2)
        nc.value = i + 1
        nc.font = font(16, True, BRAND)
        nc.alignment = align("center", "center", False)
        nc.fill = fill(BRAND_L if i % 2 == 0 else PALE)
        nc.border = thin_border

        for si, small in enumerate(smalls):
            sm = ws.cell(row=r+si, column=5)
            sm.value = small
            sm.font = font(11, False, INK)
            sm.alignment = align("left", "center", True)
            sm.fill = fill(BRAND_L if i % 2 == 0 else PALE)
            sm.border = thin_border
            sc = ws.cell(row=r+si, column=6)
            sc.font = font(13, True, BRAND_D)
            sc.alignment = align("center", "center", False)
            sc.fill = fill(WHITE)
            sc.border = medium_brand
            dv.add(sc.coordinate)
            ws.row_dimensions[r+si].height = 28
        r += 2

    end_data_row = r - 1

    # カラースケール（満足度）
    rule = ColorScaleRule(
        start_type='num', start_value=0, start_color=WARN_L,
        mid_type='num', mid_value=5, mid_color=BRAND_L,
        end_type='num', end_value=10, end_color=BRAND
    )
    ws.conditional_formatting.add(f"F{start_data_row}:F{end_data_row}", rule)

    # ── レーダーチャート（大テーマ単位の平均値） ──
    chart_start = r + 2
    ws.row_dimensions[chart_start - 1].height = 8
    # 集計テーブル（チャート用）
    ws.merge_cells(f"B{chart_start}:G{chart_start}")
    th = ws.cell(row=chart_start, column=2)
    th.value = "🎯 バランス車輪 ビジュアル化"
    th.font = font(14, True, BRAND_D)
    th.alignment = align("left", "center", False)

    chart_start += 1
    # 大テーマ名と平均値の縦リスト（隠し列：B〜C）
    ws.cell(row=chart_start, column=2).value = "テーマ"
    ws.cell(row=chart_start, column=3).value = "平均値"
    ws.cell(row=chart_start, column=2).font = font(10, True, MUTED)
    ws.cell(row=chart_start, column=3).font = font(10, True, MUTED)
    for i, (big, smalls) in enumerate(WHEEL):
        ws.cell(row=chart_start+1+i, column=2).value = big
        ws.cell(row=chart_start+1+i, column=3).value = f"=IFERROR(D{start_data_row + i*2},0)"
        ws.cell(row=chart_start+1+i, column=2).font = font(10, False, INK)
        ws.cell(row=chart_start+1+i, column=3).font = font(10, False, ACCENT)
        ws.cell(row=chart_start+1+i, column=3).number_format = "0.0"

    chart = RadarChart()
    chart.type = "filled"
    chart.style = 26
    chart.title = "人生のバランス車輪"
    labels = Reference(ws, min_col=2, min_row=chart_start+1, max_row=chart_start+8)
    data = Reference(ws, min_col=3, min_row=chart_start, max_row=chart_start+8)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.width = 16
    chart.height = 12
    ws.add_chart(chart, f"H{start_data_row}")

    # 列幅
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 4
    for c in "HIJKLMNOPQRS":
        ws.column_dimensions[c].width = 9


# ===========================================================
# 5) テーマ別 望むこと
# ===========================================================
def themes_sheet(wb):
    ws = wb.create_sheet("④テーマ別 望むこと")
    ws.sheet_view.showGridLines = False

    # タイトル帯
    ws.row_dimensions[1].height = 6
    for c in range(1, 12):
        ws.cell(row=1, column=c).fill = fill(BRAND)
    ws.row_dimensions[2].height = 38
    ws.merge_cells("B2:K2")
    t = ws["B2"]
    t.value = "テーマ別 望むこと"
    t.font = font(20, True, BRAND_D)
    t.alignment = align("left", "center", False)

    ws.row_dimensions[3].height = 36
    ws.merge_cells("B3:K3")
    d = ws["B3"]
    d.value = ("各テーマの小テーマごとに、将来「こうなっていたい」望みを書き出します。\n"
               "📌 ヒント：「すでに実現した」気持ちで、過去形・断定形で書くと効果的です。1テーマ平均6〜8個が目安。")
    d.font = font(10.5, False, MUTED)
    d.alignment = align("left", "top", True)

    # 2テーマ × 2小テーマを1ブロックとして、8テーマぶん配置
    # レイアウト：大テーマ1 (列B-E)、大テーマ2 (列G-J)
    # 1ブロックあたり 12行ほど
    r = 5
    coord_map = []  # (small_theme_label, [cell coords])
    # テーマを2列ずつのレイアウトに分割
    for block_idx in range(0, len(WHEEL), 2):
        big_left = WHEEL[block_idx]
        big_right = WHEEL[block_idx + 1] if block_idx + 1 < len(WHEEL) else None

        # 大テーマ見出し
        ws.row_dimensions[r].height = 30
        ws.merge_cells(f"B{r}:E{r}")
        lc = ws.cell(row=r, column=2)
        lc.value = big_left[0]
        lc.font = font(13, True, WHITE)
        lc.fill = fill(BRAND)
        lc.alignment = align("left", "center", False)
        lc.border = thin_border
        if big_right:
            ws.merge_cells(f"G{r}:J{r}")
            rc = ws.cell(row=r, column=7)
            rc.value = big_right[0]
            rc.font = font(13, True, WHITE)
            rc.fill = fill(BRAND)
            rc.alignment = align("left", "center", False)
            rc.border = thin_border
        r += 1

        # 小テーマ ヘッダー
        ws.row_dimensions[r].height = 24
        # 左側
        ws.cell(row=r, column=2).value = "#"
        ws.cell(row=r, column=2).fill = fill(BRAND_D)
        ws.cell(row=r, column=2).font = font(10, True, WHITE)
        ws.cell(row=r, column=2).alignment = align("center","center")
        ws.cell(row=r, column=2).border = thin_border

        ws.cell(row=r, column=3).value = big_left[1][0]
        ws.cell(row=r, column=3).fill = fill(BRAND_D)
        ws.cell(row=r, column=3).font = font(11, True, WHITE)
        ws.cell(row=r, column=3).alignment = align("center","center")
        ws.cell(row=r, column=3).border = thin_border

        ws.cell(row=r, column=4).value = "#"
        ws.cell(row=r, column=4).fill = fill(BRAND_D)
        ws.cell(row=r, column=4).font = font(10, True, WHITE)
        ws.cell(row=r, column=4).alignment = align("center","center")
        ws.cell(row=r, column=4).border = thin_border

        ws.cell(row=r, column=5).value = big_left[1][1]
        ws.cell(row=r, column=5).fill = fill(BRAND_D)
        ws.cell(row=r, column=5).font = font(11, True, WHITE)
        ws.cell(row=r, column=5).alignment = align("center","center")
        ws.cell(row=r, column=5).border = thin_border

        # 右側
        if big_right:
            ws.cell(row=r, column=7).value = "#"
            ws.cell(row=r, column=7).fill = fill(BRAND_D)
            ws.cell(row=r, column=7).font = font(10, True, WHITE)
            ws.cell(row=r, column=7).alignment = align("center","center")
            ws.cell(row=r, column=7).border = thin_border

            ws.cell(row=r, column=8).value = big_right[1][0]
            ws.cell(row=r, column=8).fill = fill(BRAND_D)
            ws.cell(row=r, column=8).font = font(11, True, WHITE)
            ws.cell(row=r, column=8).alignment = align("center","center")
            ws.cell(row=r, column=8).border = thin_border

            ws.cell(row=r, column=9).value = "#"
            ws.cell(row=r, column=9).fill = fill(BRAND_D)
            ws.cell(row=r, column=9).font = font(10, True, WHITE)
            ws.cell(row=r, column=9).alignment = align("center","center")
            ws.cell(row=r, column=9).border = thin_border

            ws.cell(row=r, column=10).value = big_right[1][1]
            ws.cell(row=r, column=10).fill = fill(BRAND_D)
            ws.cell(row=r, column=10).font = font(11, True, WHITE)
            ws.cell(row=r, column=10).alignment = align("center","center")
            ws.cell(row=r, column=10).border = thin_border
        r += 1

        # 入力欄 10行
        # それぞれの小テーマの入力セル座標を記録
        left_small1_cells = []
        left_small2_cells = []
        right_small1_cells = []
        right_small2_cells = []
        for i in range(10):
            ws.row_dimensions[r].height = 24
            # 左
            nc = ws.cell(row=r, column=2)
            nc.value = i + 1
            nc.font = font(10, False, MUTED)
            nc.alignment = align("center", "center")
            nc.fill = fill(PALE)
            nc.border = thin_border

            ic1 = ws.cell(row=r, column=3)
            ic1.font = font(11, False, INK)
            ic1.alignment = align("left", "center", True)
            ic1.fill = fill(WHITE)
            ic1.border = thin_border
            left_small1_cells.append(ic1.coordinate)

            nc2 = ws.cell(row=r, column=4)
            nc2.value = i + 1
            nc2.font = font(10, False, MUTED)
            nc2.alignment = align("center", "center")
            nc2.fill = fill(PALE)
            nc2.border = thin_border

            ic2 = ws.cell(row=r, column=5)
            ic2.font = font(11, False, INK)
            ic2.alignment = align("left", "center", True)
            ic2.fill = fill(WHITE)
            ic2.border = thin_border
            left_small2_cells.append(ic2.coordinate)

            # 右
            if big_right:
                nc3 = ws.cell(row=r, column=7)
                nc3.value = i + 1
                nc3.font = font(10, False, MUTED)
                nc3.alignment = align("center", "center")
                nc3.fill = fill(PALE)
                nc3.border = thin_border

                ic3 = ws.cell(row=r, column=8)
                ic3.font = font(11, False, INK)
                ic3.alignment = align("left", "center", True)
                ic3.fill = fill(WHITE)
                ic3.border = thin_border
                right_small1_cells.append(ic3.coordinate)

                nc4 = ws.cell(row=r, column=9)
                nc4.value = i + 1
                nc4.font = font(10, False, MUTED)
                nc4.alignment = align("center", "center")
                nc4.fill = fill(PALE)
                nc4.border = thin_border

                ic4 = ws.cell(row=r, column=10)
                ic4.font = font(11, False, INK)
                ic4.alignment = align("left", "center", True)
                ic4.fill = fill(WHITE)
                ic4.border = thin_border
                right_small2_cells.append(ic4.coordinate)
            r += 1

        # 各小テーマ × 大テーマで保存
        coord_map.append((big_left[0], big_left[1][0], left_small1_cells))
        coord_map.append((big_left[0], big_left[1][1], left_small2_cells))
        if big_right:
            coord_map.append((big_right[0], big_right[1][0], right_small1_cells))
            coord_map.append((big_right[0], big_right[1][1], right_small2_cells))

        # 区切り行
        ws.row_dimensions[r].height = 10
        r += 1

    # 列幅
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 5
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["F"].width = 4
    ws.column_dimensions["G"].width = 5
    ws.column_dimensions["H"].width = 32
    ws.column_dimensions["I"].width = 5
    ws.column_dimensions["J"].width = 32

    # 100個シート用にcoordsを返す
    return coord_map


# ===========================================================
# 6) 望むこと100個
# ===========================================================
def list100_sheet(wb):
    # ④で記録されたcoord_mapを再取得するため、再度計算する代わりに
    # ④シートと同じレイアウトロジックで参照を組み立てる
    # （ここではテーマ別シートの参照を組み立て直す）
    ws = wb.create_sheet("⑤望むこと100個")
    ws.sheet_view.showGridLines = False

    # タイトル帯
    ws.row_dimensions[1].height = 6
    for c in range(1, 8):
        ws.cell(row=1, column=c).fill = fill(BRAND)
    ws.row_dimensions[2].height = 38
    ws.merge_cells("B2:F2")
    t = ws["B2"]
    t.value = "望むこと 100個"
    t.font = font(20, True, BRAND_D)
    t.alignment = align("left", "center", False)

    ws.row_dimensions[3].height = 28
    ws.merge_cells("B3:F3")
    d = ws["B3"]
    d.value = "④テーマ別シートに入力した「望むこと」が、ここに自動で集約されます。"
    d.font = font(10.5, False, MUTED)
    d.alignment = align("left", "center", True)

    # ④シートの各小テーマブロックは
    # 行 r=5 から開始、各ブロックは：見出し1行 + 小テーマヘッダー1行 + 10行入力 + 区切り1行 = 13行
    # 1ブロックは2大テーマ（小テーマ4つ）を含む
    # 左の小テーマ1: 列C、 左の小テーマ2: 列E
    # 右の小テーマ1: 列H、 右の小テーマ2: 列J
    # データ開始行は見出し2行スキップ → r+2 から10行
    sheet_name = "④テーマ別 望むこと"

    # 100個リストを並べる列を構築
    # 各小テーマブロック→10入力セルを順に列挙
    refs = []  # (big_theme, small_theme, cell_ref)
    block_height = 13
    block_start = 5  # 最初の大テーマ見出し行
    # block_idxは大テーマペアごと → 4ブロック
    for block_idx in range(0, len(WHEEL), 2):
        big_left = WHEEL[block_idx]
        big_right = WHEEL[block_idx + 1] if block_idx + 1 < len(WHEEL) else None
        block_top = block_start + (block_idx // 2) * block_height
        data_top = block_top + 2  # 見出し + 小テーマ見出し をスキップ
        for i in range(10):
            r2 = data_top + i
            refs.append((big_left[0], big_left[1][0], f"C{r2}"))
            refs.append((big_left[0], big_left[1][1], f"E{r2}"))
            if big_right:
                refs.append((big_right[0], big_right[1][0], f"H{r2}"))
                refs.append((big_right[0], big_right[1][1], f"J{r2}"))

    # 全refsを大テーマでグループ化 → 100個のシートに大テーマ毎にまとめて並べる
    # 大テーマの順序を保ちつつ、各テーマで小テーマ別に
    grouped = {big: [] for big, _ in WHEEL}
    for big, small, cell in refs:
        grouped[big].append((small, cell))

    # ヘッダー
    r = 5
    headers = ["No.", "大テーマ", "小テーマ", "望むこと"]
    for i, h in enumerate(headers):
        c = ws.cell(row=r, column=i+2)
        c.value = h
        c.font = font(11, True, WHITE)
        c.fill = fill(BRAND)
        c.alignment = align("center", "center", False)
        c.border = thin_border
    ws.row_dimensions[r].height = 28

    r += 1
    no = 1
    for big, items in grouped.items():
        for small, cell in items:
            # No
            nc = ws.cell(row=r, column=2)
            nc.value = no
            nc.font = font(11, True, BRAND_D)
            nc.alignment = align("center", "center", False)
            nc.fill = fill(BRAND_L if no % 2 == 1 else WHITE)
            nc.border = thin_border
            # 大テーマ
            bc = ws.cell(row=r, column=3)
            bc.value = big
            bc.font = font(10.5, False, MUTED)
            bc.alignment = align("left", "center", True)
            bc.fill = fill(BRAND_L if no % 2 == 1 else WHITE)
            bc.border = thin_border
            # 小テーマ
            sc = ws.cell(row=r, column=4)
            sc.value = small
            sc.font = font(10.5, False, MUTED)
            sc.alignment = align("left", "center", True)
            sc.fill = fill(BRAND_L if no % 2 == 1 else WHITE)
            sc.border = thin_border
            # 望むこと（参照）
            tc = ws.cell(row=r, column=5)
            tc.value = f"='{sheet_name}'!{cell}"
            tc.font = font(11, False, INK)
            tc.alignment = align("left", "center", True)
            tc.fill = fill(WHITE)
            tc.border = thin_border

            ws.row_dimensions[r].height = 24
            r += 1
            no += 1
            if no > 100:
                break
        if no > 100:
            break

    # フッター
    r += 1
    ws.merge_cells(f"B{r}:E{r}")
    fc = ws.cell(row=r, column=2)
    fc.value = "🌟 すべて書き出せたら、定期的に読み返し、達成したものに印をつけていきましょう。"
    fc.font = font(11, True, BRAND_D)
    fc.alignment = align("center", "center", False)
    fc.fill = fill(BRAND_L)
    fc.border = thin_border
    ws.row_dimensions[r].height = 28

    # 列幅
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 3


if __name__ == "__main__":
    build()
